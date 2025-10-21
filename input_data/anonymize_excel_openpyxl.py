"""Anonymize Table_Genomic_Samplesv2.xlsx using openpyxl (no pandas required).

This script uses openpyxl to read and write .xlsx files directly. It loads
`.anon_mappings.json` (or creates it) and applies mappings to identifier
columns. New labels are created for unseen identifiers and saved back to the
mapping file.

Install openpyxl in your environment if missing:
  pip install openpyxl
or (conda):
  conda install -c anaconda openpyxl

Usage:
  python input_data/anonymize_excel_openpyxl.py
"""

from pathlib import Path
import json
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl (or conda install -c anaconda openpyxl)")
    sys.exit(2)

TARGET_COLUMNS = ["population", "grower", "farm", "field", "bamid", "sampleid", "group"]
PREFIXES = {
    "population": "POP",
    "grower": "GROWER",
    "farm": "FARM",
    "field": "FIELD",
    "bamid": "BAM",
    "sampleid": "SAMPLE",
    "group": "GROUP",
}

MAPPING_FILE = Path(__file__).parent / ".anon_mappings.json"
INPUT_XLSX = Path(__file__).parent / "Table_Genomic_Samplesv2.xlsx"
OUTPUT_XLSX = Path(__file__).parent / "Table_Genomic_Samplesv2_anon_openpyxl.xlsx"


def load_mappings():
    if MAPPING_FILE.exists():
        with MAPPING_FILE.open("r", encoding="utf-8") as f:
            data = json.load(f)
            for col in TARGET_COLUMNS:
                data.setdefault(col, {})
            return data
    else:
        return {col: {} for col in TARGET_COLUMNS}


def save_mappings(mappings):
    with MAPPING_FILE.open("w", encoding="utf-8") as f:
        json.dump(mappings, f, indent=2, ensure_ascii=False)


def next_label(existing_map, prefix):
    if not existing_map:
        return f"{prefix}_1"
    nums = []
    for v in existing_map.values():
        parts = str(v).rsplit("_", 1)
        if len(parts) == 2 and parts[1].isdigit():
            nums.append(int(parts[1]))
    n = max(nums) + 1 if nums else 1
    return f"{prefix}_{n}"


def anonymize_workbook(wb, mappings):
    ws = wb.active
    # read header
    header = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # map header indices to canonical lowercase
    header_lc = [str(h).lower() for h in header]

    # find indices of target columns
    target_idx = {i: header_lc[i] for i in range(len(header_lc)) if header_lc[i] in TARGET_COLUMNS}

    if not target_idx:
        print("No target identifier columns found in workbook header. No changes made.")
        return False

    # iterate rows starting from row 2
    for row in ws.iter_rows(min_row=2, values_only=False):
        for idx, key in target_idx.items():
            cell = row[idx]
            val = cell.value
            if val is None or str(val).strip() == "":
                continue
            col_map = mappings.setdefault(key, {})
            s = str(val)
            if s in col_map:
                cell.value = col_map[s]
            else:
                new_label = next_label(col_map, PREFIXES.get(key, key.upper()))
                col_map[s] = new_label
                mappings[key] = col_map
                cell.value = new_label

    return True


def main():
    if not INPUT_XLSX.exists():
        print(f"Input file not found: {INPUT_XLSX}")
        sys.exit(1)

    mappings = load_mappings()

    try:
        wb = load_workbook(INPUT_XLSX)
    except Exception as e:
        print(f"Failed to read workbook: {e}")
        sys.exit(1)

    changed = anonymize_workbook(wb, mappings)

    if changed:
        try:
            wb.save(OUTPUT_XLSX)
        except Exception as e:
            print(f"Failed to save anonymized workbook: {e}")
            sys.exit(1)
        save_mappings(mappings)
        print(f"Wrote anonymized workbook: {OUTPUT_XLSX.name}")


if __name__ == "__main__":
    main()
