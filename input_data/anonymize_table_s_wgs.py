"""Anonymize the 'Table_S#_WGS_Samples' sheet in Table_Genomic_Samplesv2.xlsx.

This script uses openpyxl and the existing mapping file `input_data/.anon_mappings.json`
to replace identifiers in the specified sheet. It targets the columns: sample,
population and group (case-insensitive). If new identifiers are encountered the
script will append new anonymized labels to the mapping file.

Run this using your conda environment `jfmr` so dependencies (openpyxl) are
available. Example:

  conda run -n jfmr python input_data/anonymize_table_s_wgs.py

The script writes an output workbook named
`Table_Genomic_Samplesv2_Table_S#_WGS_Samples_anon.xlsx` alongside the input.
"""

from pathlib import Path
import json
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not found. Run this script inside your jfmr environment: conda run -n jfmr python <script>")
    sys.exit(2)

MAPPING_FILE = Path(__file__).parent / ".anon_mappings.json"
INPUT_XLSX = Path(__file__).parent / "Table_Genomic_Samplesv2.xlsx"
SHEET_NAME = "Table_S#_WGS_Samples"
OUTPUT_XLSX = Path(__file__).parent / f"Table_Genomic_Samplesv2_{SHEET_NAME}_anon.xlsx"

# canonical target columns (lowercase)
TARGET_COLUMNS = ["sample", "population", "group"]
PREFIXES = {
    "sample": "SAMPLE",
    "population": "POP",
    "group": "GROUP",
}


def load_mappings():
    if MAPPING_FILE.exists():
        with MAPPING_FILE.open("r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {}
    # ensure target keys exist
    for k in TARGET_COLUMNS:
        data.setdefault(k, {})
    return data


def save_mappings(mappings):
    with MAPPING_FILE.open("w", encoding="utf-8") as f:
        json.dump(mappings, f, indent=2, ensure_ascii=False)


def next_label(existing_map, prefix):
    if not existing_map:
        return f"{prefix}_1"
    nums = []
    for v in existing_map.values():
        parts = str(v).rsplit("_", 1)
        if len(parts) == 2 and parts[1].isdigit():
            nums.append(int(parts[1]))
    n = max(nums) + 1 if nums else 1
    return f"{prefix}_{n}"


def anonymize_sheet(wb, sheet_name, mappings):
    if sheet_name not in wb.sheetnames:
        raise FileNotFoundError(f"Sheet not found: {sheet_name}")

    ws = wb[sheet_name]
    # header row
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    header = [cell.value if cell.value is not None else "" for cell in header_cells]
    header_lc = [str(h).lower() for h in header]

    # map positions for the target columns (case-insensitive)
    target_positions = {i: header_lc[i] for i in range(len(header_lc)) if header_lc[i] in TARGET_COLUMNS}

    if not target_positions:
        print("No target columns found in sheet header. Nothing to do.")
        return False

    for row in ws.iter_rows(min_row=2, values_only=False):
        for idx, key in target_positions.items():
            cell = row[idx]
            val = cell.value
            if val is None or str(val).strip() == "":
                continue
            lc = key  # canonical lowercase key
            col_map = mappings.setdefault(lc, {})
            s = str(val)
            if s in col_map:
                cell.value = col_map[s]
            else:
                prefix = PREFIXES.get(lc, lc.upper())
                new_label = next_label(col_map, prefix)
                col_map[s] = new_label
                mappings[lc] = col_map
                cell.value = new_label

    return True


def main():
    if not INPUT_XLSX.exists():
        print(f"Input file not found: {INPUT_XLSX}")
        sys.exit(1)

    mappings = load_mappings()

    try:
        wb = load_workbook(INPUT_XLSX)
    except Exception as e:
        print(f"Failed to open workbook: {e}")
        sys.exit(1)

    try:
        changed = anonymize_sheet(wb, SHEET_NAME, mappings)
    except FileNotFoundError as e:
        print(e)
        sys.exit(1)

    if changed:
        try:
            wb.save(OUTPUT_XLSX)
        except Exception as e:
            print(f"Failed to save anonymized workbook: {e}")
            sys.exit(1)
        save_mappings(mappings)
        print(f"Wrote anonymized workbook: {OUTPUT_XLSX.name}")
    else:
        print("No changes made.")


if __name__ == "__main__":
    main()
